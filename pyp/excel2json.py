import sys
import os
import json
from openpyxl import load_workbook
from operator import attrgetter

cwd = os.getcwd()

def toCamelStr(str):
    camelStr = ''
    charIdx = 0
    upperIdx = -1
    for char1 in str:
        if char1 == '_':
            upperIdx = charIdx+1
        elif charIdx == upperIdx:
            camelStr = camelStr + char1.upper()
        else:
            camelStr = camelStr + char1
        charIdx+=1
    return camelStr
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('param should be: excel2json xxx')
        exit(1)
    excelFile = sys.argv[1]
    wb = load_workbook("data2/" + excelFile + ".xlsx", data_only=True)
    sheetName = wb.sheetnames[0]
    # print(sheetName)
    dataSheet = wb[sheetName]
    maxRow = dataSheet.max_row
    # print(dataSheet, maxRow)
    sheetAttrs = dataSheet["2"]

    attrListJson = json.load(open("data2/client_config_list.json", encoding='UTF-8'))  ##读取属性列表
    notCamelUseAttrs = attrListJson[sheetName]
    if notCamelUseAttrs is None:
        print('cannot find sheetName!')
        exit(1)
    notCamelUseAttrs = sorted(notCamelUseAttrs)
    useAttrs = []
    cIdx = 0
    for nc in notCamelUseAttrs:
        useAttrs.append('')
        for char1 in nc:
            if char1.isupper():
                useAttrs[cIdx] = useAttrs[cIdx] + '_' + char1.lower()
            else:
                useAttrs[cIdx] = useAttrs[cIdx] + char1
        for isIdx in range(len(list(sheetAttrs))):
            blk = list(sheetAttrs)[isIdx]
            if blk.value == useAttrs[cIdx]:
                useAttrs[cIdx] = blk
                break
            elif isIdx == len(list(sheetAttrs)) -1:
                print(str(useAttrs[cIdx]) + ' not in sheet attrs!')
                exit(1)
        cIdx = cIdx + 1
    
    # for aa in useAttrs:
    #     print(aa.value)
    # exit(0)

    # useAttrs = []
    # for i in list(sheetAttrs):
    #     if i.value is not None:
    #         # print(i)
    #         useAttrs.append(i)
    # useAttrs = sorted(useAttrs, key=attrgetter('value'))
    # print("sorted=")  # 加上客户端过滤字段
    # for j in useAttrs:
    #     print(j.column, j.coordinate, j.row, j.value)  # chr(64+j.column),
    
    datatype = attrListJson[sheetName+'_datatype']
    jfpath = "data2/" + excelFile + ".json"
    jf = open(jfpath,'w')
    jf.writelines('\n[\n')
    for k in range(3, maxRow):
        if k > 3:
            jf.writelines(',\n\t\n\t{')
        else:
            jf.writelines('\t\n\t{')
        j2Idx = 0
        for j2 in useAttrs:
            col = j2.column
            if str(j2.column).isdigit():  # 坑爹的偶尔会读出来是数字
                col = chr(64+j2.column)
            cell = dataSheet[str(col) + str(k)]  # chr(64+j2.column) + str(k)
            if cell.value is not None:
                prefix = ',\n\t\t"' if j2Idx > 0 else '\n\t\t"'
                attrName = toCamelStr(j2.value)
                cellValue = str(cell.value)
                if datatype[attrName] == 2: ##字符串类型
                    if '"' in cellValue:
                        cellValue = cellValue.replace('"', '\\"', -1)
                    cellValue = '"' + cellValue + '"'
                jf.writelines(prefix + attrName + '" : ' + cellValue + '')
                j2Idx+=1
                # print(col, cell.value)
        jf.writelines('\n\t}')
        print('--------------')
    jf.writelines('\n]\n')
    jf.close()
    # 组装成json-ok
    # 打包config.json
